from simulations.simulation import *
from simulations.plots import *
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill, Border, Side
from os import startfile, path

def run_full_simulation(df, player, n_sim=1000, inactive_players=None):
    today = df['date'].max()
    schedule = build_future_schedule(today + pd.Timedelta(days=1), '2026-04-30')
    profiles = build_player_profiles(df)

    real_cutoff = compute_real_cutoff(df)
    real_player = get_real_player_path(df, player)

    all_cutoffs, all_players = run_simulations(df, profiles, schedule, n_sim, inactive_players=inactive_players)
    expected_cutoff = compute_expected_cutoff(all_cutoffs)
    expected_player = compute_expected_player_path(all_players, player)

    sim_cutoff = [
        {'date': d, 'cutoff': v}
        for d, v in expected_cutoff.items()
    ]
    sim_player = [
        {'date': d, 'points': v}
        for d, v in expected_player.items()
    ]

    return {
        'real_cutoff': real_cutoff,
        'real_player': real_player,
        'sim_cutoff': sim_cutoff,
        'sim_player': sim_player,
        'all_players': all_players
    }

def save_rank_projection(ranking, filepath='data/expected_final_ranking.txt'):
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('EXPECTED FINAL RANKING (FROM SIMULATION)\n')
        f.write('=' * 50 + '\n\n')

        for i, (player, avg_rank) in enumerate(ranking, 1):
            f.write(f'{i}. {player} - avg rank: {avg_rank:.2f}\n')
    print('Expected Final Ranking saved to data/expected_final_ranking.txt')

def run_rank_projection_pipeline(df, all_players, today, output_top, save=True, eval_pool=None):
    standings = get_current_standings(df)
    sorted_players = sorted(standings.items(), key=lambda x: x[1], reverse=True)
    if eval_pool:
        selected_players = [p for p, _ in sorted_players[:eval_pool]]
    else:
        selected_players = [p for p, _ in sorted_players]

    real_ranks_dict = {}
    sim_ranks_dict = {}

    for player in selected_players:
        real_ranks_dict[player] = get_real_player_rank_path(df, player)
        sim_ranks_dict[player] = compute_expected_player_rank(all_players, player)

    plot_rank_projections_multi(real_ranks_dict, sim_ranks_dict, today, top_n=output_top)
    final_ranking = compute_expected_final_ranking(all_players, eval_pool, output_top)
    if save:
        save_rank_projection(final_ranking)

def save_playoff_odds_csv(df, path='data/playoff_odds.csv'):
    df.to_csv(path, encoding='utf-8')

def save_playoff_odds_excel(df, path='data/playoff_odds.xlsx'):
    top_col = 'Top 18 Prob'
    cols = [top_col] + [c for c in df.columns if c != top_col]
    df = df[cols]
    df = df.copy()
    df['Expected Rank'] = 0
    for i in range(1, 20):
        col = f'Rank {i}'
        if col in df.columns:
            df['Expected Rank'] += df[col] * i
    df = df.sort_values(by=[top_col, 'Expected Rank'], ascending=[False, True])
    df = df.drop(columns=['Expected Rank'])

    df.insert(1, 'Final Rank', range(1, len(df) + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Playoff Odds'
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        ws.append(row)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if isinstance(cell.value, float):
                if cell.value == 0:
                    cell.value = ''
                else:
                    cell.number_format = '0.0%'
    max_row = ws.max_row
    max_col = ws.max_column

    top18_range = f'{ws.cell(row=2, column=2).coordinate}:{ws.cell(row=max_row, column=2).coordinate}'
    rank_prob_range = f'{ws.cell(row=2, column=4).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}'
    gradient_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='FF0000',
        mid_type='percentile', mid_value=50, mid_color='FFFF00',
        end_type='num', end_value=1, end_color='00FF00'
    )
    ws.conditional_formatting.add(top18_range, gradient_rule)
    ws.conditional_formatting.add(rank_prob_range, gradient_rule)

    rank_col = 3
    rank_range = f'{ws.cell(row=3, column=rank_col).coordinate}:{ws.cell(row=max_row, column=rank_col).coordinate}'
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    top18_rule = CellIsRule(
        operator='lessThanOrEqual',
        formula=['19'],
        fill=green_fill
    )
    ws.conditional_formatting.add(rank_range, top18_rule)

    red_line = Side(style='thick', color='FF0000')
    border = Border(bottom=red_line)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=21, column=col)
        cell.border = border

    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'B2'
    wb.save(path)

def run_playoff_odds_pipeline(all_players, output_top=19, eval_pool=50, save_csv=True, save_excel=True, open_excel=True):
    odds_df = compute_playoff_odds(all_players, output_top, eval_pool)
    if save_csv:
        save_playoff_odds_csv(odds_df)
        print('Playoff Odds saved to data/playoff_odds.csv')
    if save_excel:
        save_playoff_odds_excel(odds_df)
        print('Playoff Odds saved to data/playoff_odds.xlsx')
    if open_excel:
        startfile(path.abspath(path.join('data', 'playoff_odds.xlsx')))
    return odds_df